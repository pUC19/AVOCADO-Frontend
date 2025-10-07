# app.py — POTATO Flask Frontend (Bootstrap + Templates)
# Fortschrittsbalken (spawn-fest) + robuste Direction-Erkennung (unfolding/refolding)
from batch_report import report_bp
import os, io, json, zipfile, re
import csv  # added for csv.Sniffer
import seaborn as sns
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.dataframe import dataframe_to_rows

from pathlib import Path
from datetime import datetime
from multiprocessing import Process, Queue
from queue import Empty
import uuid
import math

from flask import (
    Flask, request, redirect, url_for, send_file,
    render_template, abort, jsonify, flash
)

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

# POTATO-Backend
from POTATO_ForceRamp import read_in_data, start_subprocess
from POTATO_config import (
    default_values_HF, default_values_LF, default_values_CSV, default_values_FIT
)
import POTATO_find_steps as PFS
from POTATO_fitting import fitting_ds, fitting_ss

# Import the function to add the Dash app
from dna_analyzer import add_dash_to_flask

# --- matplotlib fill_between guard (verhindert Length-Mismatch-Fehler) ---
import numpy as _np
import matplotlib.axes as _maxes
try:
    _orig_fill_between
except NameError:
    _orig_fill_between = _maxes.Axes.fill_between

def _fill_between_guard(self, x, y1, y2=0, **kw):
    x  = _np.asarray(x); y1 = _np.asarray(y1)
    y2a = _np.asarray(y2) if _np.ndim(y2) else _np.full_like(x[:len(y1)], y2, dtype=float)
    n = min(len(x), len(y1), len(y2a))
    if n <= 0: return
    return _orig_fill_between(self, x[:n], y1[:n], y2a[:n], **kw)

_maxes.Axes.fill_between = _fill_between_guard
# ------------------------------------------------------------------------

# --------------------------------------------------------------------
# App / Verzeichnisse
# --------------------------------------------------------------------
app = Flask(__name__)
app.secret_key = "change-me"
app.register_blueprint(report_bp)

# Add the Dash app to the Flask server
add_dash_to_flask(app)

BASE = Path(__file__).parent.resolve()
UPLOAD_DIR = BASE / "uploads"
ANALYSIS_DIR = BASE / "analysis"
UPLOAD_DIR.mkdir(exist_ok=True)
ANALYSIS_DIR.mkdir(exist_ok=True)

# --------------------------------------------------------------------
# Physik-Konstanten & eFJC (für ΔLc aus Step-Länge)
# --------------------------------------------------------------------
KBT_PN_NM       = 4.114    # pN·nm bei ~298 K
KUHN_B_NM       = 1.9      # nm (≈ 2*Lp_ss)
STRETCH_S_PN    = 1000.0   # pN
DEFAULT_F_REF   = 15.0     # pN – feste Referenzkraft für ΔLc
MIN_VALID_FORCE = 8.0      # pN – zu klein → ΔLc unsicher → NaN

def phi_ss_eFJC(F_pN: float,
                b_nm: float = KUHN_B_NM,
                S_pN: float = STRETCH_S_PN,
                kBT: float = KBT_PN_NM) -> float:
    if F_pN is None or not (F_pN > 0):
        return float('nan')
    y = F_pN * b_nm / kBT
    return (1.0 / math.tanh(y) - 1.0 / y) + (F_pN / S_pN)

def delta_lc_from_step(delta_x_nm: float, F_eval_pN: float) -> float:
    if delta_x_nm is None or np.isnan(delta_x_nm):
        return np.nan
    phi = phi_ss_eFJC(F_eval_pN)
    return (abs(delta_x_nm) / phi) if (phi and not math.isnan(phi)) else np.nan

# --------------------------------------------------------------------
# Monkey-Patch für POTATO-Plot (robustes fill_between)
# --------------------------------------------------------------------
def _safe_fill_between(ax, x, y1, y2, **kw):
    x = np.asarray(x); y1 = np.asarray(y1); y2 = np.asarray(y2)
    n = min(len(x), len(y1), len(y2))
    if n <= 0: return
    ax.fill_between(x[:n], y1[:n], y2[:n], **kw)

def save_figure_safe(export_PLOT, timestamp, filename_i, analysis_folder,
                     Force_Distance, derivative_array, F_trimmed, PD_trimmed,
                     PD_start_F, PD_start_PD):
    import POTATO_find_steps as _P
    y_vector_PD = getattr(_P, "y_vector_PD", [])
    y_vector_F  = getattr(_P, "y_vector_F", [])
    PD_pos = getattr(_P, "PD_mm2_STD2_positive", [])
    PD_neg = getattr(_P, "PD_mm2_STD2_negative", [])
    F_pos  = getattr(_P, "F_mm2_STD2_positive", [])
    F_neg  = getattr(_P, "F_mm2_STD2_negative", [])

    fig, axs = plt.subplots(2, 2, figsize=(10, 6), dpi=100)
    ax1, ax2, ax3, ax4 = axs.flat

    ax1.set_title("FD-Curve"); ax1.set_ylabel("Force (pN)")
    ax1.scatter(Force_Distance[:, 1], Force_Distance[:, 0], s=0.6, marker='.', alpha=1)

    ax2.set_title("Trimmed FD-Curve - steps marked")
    ax2.scatter(PD_trimmed, F_trimmed, s=0.6, marker='.', alpha=1)
    for x in PD_start_F:  ax2.axvline(x=x, color='red',   lw=0.5, alpha=0.5)
    for x in PD_start_PD: ax2.axvline(x=x, color='green', lw=0.5, alpha=0.5)

    ax3.set_title("Distance derivative"); ax3.set_xlabel("Distance (nm)"); ax3.set_ylabel("ΔDistance (nm/ms)")
    ax3.plot(derivative_array[:, 1], derivative_array[:, 3])
    if len(y_vector_PD):
        _safe_fill_between(ax3, derivative_array[:, 1], PD_pos, PD_neg, color='black', alpha=0.30)
        n = min(len(derivative_array[:, 1]), len(y_vector_PD))
        ax3.plot(derivative_array[:n, 1], np.asarray(y_vector_PD)[:n])

    ax4.set_title("Force derivative"); ax4.set_xlabel("Distance (nm)"); ax4.set_ylabel("ΔForce (pN/ms)")
    ax4.plot(derivative_array[:, 1], derivative_array[:, 2])
    if len(y_vector_F):
        _safe_fill_between(ax4, derivative_array[:, 1], F_pos, F_neg, color='black', alpha=0.30)
        n = min(len(derivative_array[:, 1]), len(y_vector_F))
        ax4.plot(derivative_array[:n, 1], np.asarray(y_vector_F)[:n])

    fig.tight_layout()
    if export_PLOT == 1:
        out = os.path.join(analysis_folder, f"{filename_i}_plot_{timestamp}.png")
        fig.savefig(out, dpi=600)
        plt.close(fig)
    return fig

# Patch aktivieren (Hauptprozess)
PFS.save_figure = save_figure_safe

# --------------------------------------------------------------------
# Helper (Einstellungen)
# --------------------------------------------------------------------
def make_input_settings(default_block):
    return {
        'downsample_value': int(default_block['Downsampling rate']),
        'filter_degree': int(default_block['Butterworth filter degree']),
        'filter_cut_off': float(default_block['Cut-off frequency']),
        'F_min': float(default_block['Force threshold, pN']),
        'step_d': int(default_block['Step d']),
        'z-score_f': float(default_block['Z-score force']),
        'z-score_d': float(default_block['Z-score distance']),
        'window_size': int(default_block['Moving median window size']),
        'data_frequency': float(default_block['Data frequency, Hz']),
        'STD_diff': float(default_block['STD difference threshold']),
        'augment_factor': ''
    }

def make_input_format(kind="HF", trap1=True, length_um=True, preprocess=True, multi=False, augment=False):
    return {
        'HF': 1 if kind == "HF" else 0,
        'LF': 1 if kind == "LF" else 0,
        'CSV': 1 if kind == "CSV" else 0,
        'Augment': 1 if augment else 0,
        'Trap': 1 if trap1 else 0,
        'length_measure': 1 if length_um else 0,
        'MultiH5': 1 if multi else 0,
        'preprocess': 1 if preprocess else 0
    }

def make_input_fitting(default_fit=default_values_FIT, model="WLC+WLC"):
    return {
        'WLC+WLC': 1 if model == "WLC+WLC" else 0,
        'WLC+FJC': 1 if model == "WLC+FJC" else 0,
        'lp_ds': float(default_fit['Persistance-Length ds, nm']),
        'lp_ds_up': float(default_fit['Persistance-Length ds, upper bound, nm']),
        'lp_ds_low': float(default_fit['Persistance-Length ds, lower bound, nm']),
        'lc_ds': float(default_fit['Contour-Length ds, nm']),
        'lp_ss': float(default_fit['Persistance-Length ss, nm']),
        'lc_ss': float(default_fit['Contour-Length ss, nm']),
        'lc_ss_up': float(default_fit['Contour-Length ss, upper bound, nm']),
        'ds_stiff': float(default_fit['Stiffness ds, pN']),
        'ds_stiff_up': float(default_fit['Stiffness ds, upper bound, pN']),
        'ds_stiff_low': float(default_fit['Stiffness ds, lower bound, pN']),
        'ss_stiff': float(default_fit['Stiffness ss, pN']),
        'ss_stiff_up': float(default_fit['Stiffness ss, upper bound, pN']),
        'ss_stiff_low': float(default_fit['Stiffness ss, lower bound, pN']),
        'offset_f': float(default_fit['Force offset, pN']),
        'offset_f_up': float(default_fit['Force offset, upper bound, pN']),
        'offset_f_low': float(default_fit['Force offset, lower bound, pN']),
        'offset_d': float(default_fit['Distance offset, nm']),
        'offset_d_up': float(default_fit['Distance offset, upper bound, nm']),
        'offset_d_low': float(default_fit['Distance offset, lower bound, nm']),
    }

def make_export_flags(export_smooth=True, export_plot=True, export_steps=True, export_total=True, export_fit=True):
    return {
        'export_SMOOTH': 1 if export_smooth else 0,
        'export_PLOT': 1 if export_plot else 0,
        'export_STEPS': 1 if export_steps else 0,
        'export_TOTAL': 1 if export_total else 0,
        'export_FIT': 1 if export_fit else 0,
    }

def kdeplot_filled(*, data, label, color=None, lw=2, alpha=0.5):
    try:
        return sns.kdeplot(data=data, fill=True, label=label, lw=lw, alpha=alpha, color=color)
    except TypeError:
        return sns.kdeplot(data=data, shade=True, label=label, lw=lw, alpha=alpha, color=color)

# --------------------------------------------------------------------
# Direction-Erkennung (robust)
# --------------------------------------------------------------------
_UNFOLD_PATTERNS = [
    r'\b1f\b', r'\bfw\b', r'\bforward\b', r'\bunfold(?:ing)?\b', r'\bunf\b',
    r'(?<![a-z])unfold(?:ing)?(?![a-z])',
    r'(?<![a-z])forward(?![a-z])',
    r'(?<![a-z])fw(?![a-z])',
    r'(?<![a-z])unf(?![a-z])',
    # Token "f" allein (z. B. "_f", " f", "-f", "(f)") am Wortende
    r'(?<![a-z0-9])f(?![a-z0-9])',
    # Muster wie "1-1f", "2f", "3-2f" am Segment-/Dateiende
    r'(?<![a-z])\d+(?:[-_]\d+)?f(?![a-z0-9])',
]
_REFOLD_PATTERNS = [
    r'\b1r\b', r'\brv\b', r'\breverse\b', r'\brefold(?:ing)?\b', r'\bref\b',
    r'(?<![a-z])refold(?:ing)?(?![a-z])',
    r'(?<![a-z])reverse(?![a-z])',
    r'(?<![a-z])rv(?![a-z])',
    r'(?<![a-z])ref(?![a-z])',
    # Token "r" allein (z. B. "_r", " r", "-r", "(r)") am Wortende
    r'(?<![a-z0-9])r(?![a-z0-9])',
    # Muster wie "1-1r", "2r", "3-2r" am Segment-/Dateiende
    r'(?<![a-z])\d+(?:[-_]\d+)?r(?![a-z0-9])',
]

def _detect_direction_from_name(text: str) -> str:
    s = (text or "").lower()
    for pat in _UNFOLD_PATTERNS:
        if re.search(pat, s): return "unfolding"
    for pat in _REFOLD_PATTERNS:
        if re.search(pat, s): return "refolding"
    return "unknown"

def _direction_from_filename_or_path(filename: str, path: str) -> str:
    # kombiniere normalisierten Dateinamen, echten Dateinamen und Ordnernamen
    parts = [str(filename or "")]
    if path:
        p = Path(path)
        parts += [p.name, p.parent.name]
    return _detect_direction_from_name(" ".join(parts))

# --------------------------------------------------------------------
# State
# --------------------------------------------------------------------
TOMATO = {}         # token -> {filename, path, FD, freq, settings, format, fitcfg, last_steps, last_results}
JOBS = {}           # Batch-Analyse Jobs (OT)
LAST_SERVER_FOLDER = None
TOMATO_JOBS = {}    # manuelle Analyse als Job mit Fortschritt

def drain_queue(q):
    out = []
    while True:
        try: msg = q.get_nowait()
        except Empty: break
        else: out.append(str(msg))
    return out

def _drain_tomato_job(job):
    logs = drain_queue(job["queue"])
    progress = []
    for m in logs:
        if m == "__DONE__":
            job["done"] = True
        elif m.startswith("__ERROR__"):
            job["error"] = m.replace("__ERROR__:", "")
            job["done"] = True
        elif m.startswith("__PROGRESS__"):
            try:
                _, pct_str, msg = m.split(":", 2)
                progress.append({"pct": int(pct_str), "msg": msg})
            except Exception:
                pass
        elif m.startswith("__RESULT__"):
            try:
                _, payload = m.split(":", 1)
                job["payload"] = json.loads(payload)
            except Exception:
                job["error"] = "Result parsing failed"
                job["done"] = True
        else:
            pass
    return progress

# --------------------------------------------------------------------
# Datei-Liste Server
# --------------------------------------------------------------------
def list_server_files(folder: str):
    p = Path(folder)
    if not p.exists(): return []
    return sorted([str(f) for f in p.rglob("*") if f.suffix.lower() in (".h5", ".csv")])

# --------------------------------------------------------------------
# OT-Run Auswertung (Excel → ΔLc + Plots → neues Excel)
# --------------------------------------------------------------------

def _find_smooth_csvs(base_dir: str):
    """
    Suche NUR in genau base_dir nach CSV-Dateien (nicht rekursiv).
    Erwartung: base_dir ist dein 'smooth' Ordner.
    """
    p = Path(base_dir)
    if not p.exists() or not p.is_dir():
        return []
    out = []
    for f in p.iterdir():
        if f.is_file() and f.suffix.lower() == ".csv":
            out.append(str(f))
    return sorted(out)



def _read_fd_from_csv(csv_path: str):
    """
    1) Versuche TOMATO-Parser (read_in_data) – identische Skalierung/Sortierung.
    2) Fallback: robustes, headerloses Einlesen mit Sniffer & Heuristiken.
    Rückgabe: (distance_nm, force_pN) oder (None, None)
    """
    # --- 1) TOMATO-Parser ---
    try:
        input_settings, input_format, _, _ = _default_settings(kind="CSV", trap1=True, len_um=True)
        FD, FD_um, freq, filename = read_in_data(0, [str(csv_path)], input_settings, input_format)
        x = FD[:, 1].astype(float)   # Distance (nm)
        y = FD[:, 0].astype(float)   # Force (pN)
        if x.size == 0:
            return None, None
        if x[0] > x[-1]:
            x = x[::-1]; y = y[::-1]
        # Downsampling
        if x.size > 20000:
            step = max(1, x.size // 20000)
            x = x[::step]; y = y[::step]
        return x, y
    except Exception:
        pass  # → Fallback

    # --- 2) Fallback: Sniffer + Heuristiken (ohne Kopfzeile) ---
    try:
        with open(csv_path, "r", errors="ignore", newline="") as fh:
            sample = fh.read(4096)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t ")
            sep = dialect.delimiter
        except Exception:
            sep = max([",",";","\t"," "], key=lambda d: sample.count(d))
        decimal = "," if (sep == ";" and re.search(r"\d,\d", sample)) else "."
        dfc = pd.read_csv(csv_path, header=None, sep=sep, decimal=decimal)
    except Exception:
        return None, None

    # numerische Spalten
    num_cols = []
    for c in dfc.columns:
        col = pd.to_numeric(dfc[c], errors="coerce")
        if col.notna().sum() > 0:
            num_cols.append(c)
    if len(num_cols) < 2:
        return None, None
    dfc = dfc[num_cols].copy()

    def _is_time_like(v):
        v = pd.to_numeric(v, errors="coerce").to_numpy()
        v = v[np.isfinite(v)]
        if v.size < 10: return False
        if not np.all(np.diff(v) >= -1e-12): return False
        d = np.diff(v)
        if d.size < 5: return False
        mean_step = np.nanmean(d); var_step = np.nanvar(d)
        return (mean_step > 0) and (var_step / (mean_step**2 + 1e-12) < 0.05)

    cand = [c for c in dfc.columns if not _is_time_like(dfc[c])]
    if len(cand) < 2:
        cand = sorted(dfc.columns, key=lambda c: pd.to_numeric(dfc[c], errors="coerce").var(), reverse=True)[:2]

    a, b = cand[:2]
    A = pd.to_numeric(dfc[a], errors="coerce").to_numpy()
    B = pd.to_numeric(dfc[b], errors="coerce").to_numpy()

    def _span(x):
        x = x[np.isfinite(x)]
        if x.size == 0: return -np.inf
        lo, hi = np.percentile(x, [5, 95])
        return float(hi - lo)

    spanA, spanB = _span(A), _span(B)
    dist_raw, force_raw = (A, B) if spanA >= spanB else (B, A)

    def _looks_like_force(x):
        x = x[np.isfinite(x)]
        if x.size == 0: return 0
        vmin, vmax = np.nanmin(x), np.nanmax(x); rng = vmax - vmin
        score = 0
        if -20 <= vmin and vmax <= 300: score += 1
        if rng < 400: score += 0.5
        return score
    if _looks_like_force(dist_raw) > _looks_like_force(force_raw):
        dist_raw, force_raw = force_raw, dist_raw

    finite_dist = dist_raw[np.isfinite(dist_raw)]
    if finite_dist.size:
        dmin, dmax = np.nanmin(finite_dist), np.nanmax(finite_dist)
        if (dmax - dmin) < 80 and dmax < 80:  # vermutlich µm
            dist_raw = dist_raw * 1000.0

    m = np.isfinite(dist_raw) & np.isfinite(force_raw)
    x, y = dist_raw[m], force_raw[m]
    if x.size == 0: return None, None
    if x[0] > x[-1]:
        x = x[::-1]; y = y[::-1]
    if x.size > 20000:
        step = max(1, x.size // 20000)
        x = x[::step]; y = y[::step]
    return x, y

def _run_ot_auswertung_on_xlsx(xlsx_path: str) -> io.BytesIO:
    df = pd.read_excel(xlsx_path)

    if 'Filename' in df.columns:
        df = df[df['Filename'].notna()].copy()

    # Falls Direction fehlt: aus dem (Datei)Namen ermitteln
    if 'Direction' not in df.columns:
        def _detect_dir_row(s: str) -> str:
            return _detect_direction_from_name(str(s))
        df['Direction'] = df.get('Filename', '').astype(str).apply(_detect_dir_row)

    step_len_col = "Step length [nm]"
    force_col    = "mean force [pN]"
    dir_col      = "Direction"
    df["Delta Lc (nm)"] = np.nan  # signed

    if step_len_col in df.columns:
        dx  = pd.to_numeric(df[step_len_col], errors="coerce")
        Fm  = pd.to_numeric(df.get(force_col, np.nan), errors="coerce")
        sign = pd.Series(1.0, index=df.index, dtype="float64")
        if dir_col in df.columns:
            d = df[dir_col].astype(str).str.lower()
            sign[d.str.startswith("refold")] = -1.0
        out = []
        for i in df.index:
            F_eval = DEFAULT_F_REF
            Fm_i = Fm.loc[i] if i in Fm.index else np.nan
            if not np.isnan(Fm_i) and Fm_i < MIN_VALID_FORCE:
                out.append(np.nan); continue
            dlc = delta_lc_from_step(dx.loc[i], F_eval)
            if not np.isnan(dlc): dlc *= sign.loc[i]
            out.append(dlc)
        df["Delta Lc (nm)"] = out
    else:
        lc_col = 'ss contour Length'
        step_col = 'step number' if 'step number' in df.columns else ('step' if 'step' in df.columns else None)
        if lc_col in df.columns:
            for _, g in df.groupby('Filename', dropna=False, sort=False):
                g_sorted = g.sort_values(step_col) if (step_col and step_col in g.columns) else g
                lc = pd.to_numeric(g_sorted[lc_col], errors='coerce')
                df.loc[g_sorted.index, 'Delta Lc (nm)'] = lc.diff().values

    df["Delta Lc abs (nm)"] = pd.to_numeric(df["Delta Lc (nm)"], errors='coerce').abs()

    # --- Plot-Sammellisten ---
    plotfiles = []        # nur Dateipfade
    debug_lines = []      # kurze Texte fürs "Plots"-Sheet
    smooth_debug_rows = [ # ausführlicher Report fürs eigenes Sheet
        ["file", "status", "direction", "n_points", "note"]
    ]

    def _save(fig, fname):
        fig.tight_layout()
        fig.savefig(fname, bbox_inches="tight", dpi=200)
        plt.close(fig)
        plotfiles.append(fname)



    # KDE Force
    fig = plt.figure(figsize=(6,4), dpi=120)
    for direction, color, label in zip(['refolding','unfolding'], ['#5dade2','#f5b041'], ['refolding','unfolding']):
        data = pd.to_numeric(df.loc[df['Direction']==direction, force_col], errors='coerce').dropna()
        if not data.empty: kdeplot_filled(data=data, label=label, color=color, lw=2, alpha=0.5)
    plt.xlabel("Force (pN)"); plt.ylabel("Wahrscheinlichkeit"); plt.title("PDF: Force (pN)"); plt.legend()
    _save(fig, "plot_pdf_force.png")

    # KDE |ΔLc|
    fig = plt.figure(figsize=(6,4), dpi=120)
    for direction, color, label in zip(['refolding','unfolding'], ['#5dade2','#f5b041'], ['refolding','unfolding']):
        data = pd.to_numeric(df.loc[df['Direction']==direction, "Delta Lc abs (nm)"], errors='coerce').dropna()
        if not data.empty: kdeplot_filled(data=data, label=label, color=color, lw=2, alpha=0.5)
    plt.xlabel("|ΔLc| (nm)"); plt.ylabel("Wahrscheinlichkeit"); plt.title("PDF: |ΔLc| (nm)"); plt.legend()
    _save(fig, "plot_pdf_deltalc_abs.png")

    # HEXBIN helper
    def _hexbin_2panel(x_series, y_series, title, xlab, ylab, gridsize=35):
        x_all = pd.to_numeric(x_series, errors='coerce')
        y_all = pd.to_numeric(y_series, errors='coerce')
        mask  = (~x_all.isna()) & (~y_all.isna())
        if not mask.any(): return None
        xmin, xmax = float(x_all[mask].min()), float(x_all[mask].max())
        ymin, ymax = float(y_all[mask].min()), float(y_all[mask].max())
        extent = (xmin, xmax, ymin, ymax)
        fig, axs = plt.subplots(1, 2, figsize=(11,4.5), dpi=120, sharex=True, sharey=True)
        fig.suptitle(title)
        for ax, direction, subtitle in zip(axs, ['unfolding','refolding'], ['Unfolding','Refolding']):
            sel = (df['Direction']==direction)
            x = pd.to_numeric(x_series[sel], errors='coerce')
            y = pd.to_numeric(y_series[sel], errors='coerce')
            m = (~x.isna()) & (~y.isna())
            if m.any():
                hb = ax.hexbin(x[m], y[m], gridsize=gridsize, extent=extent, bins='log')
                cb = fig.colorbar(hb, ax=ax); cb.set_label("Anzahl (log)")
            ax.set_title(subtitle); ax.set_xlabel(xlab); ax.set_ylabel(ylab)
        return fig

    fig = _hexbin_2panel(df["Delta Lc abs (nm)"], df[force_col],
                         "Unfolding/Refolding – Fc vs |ΔLc| (Hexbin)",
                         "|ΔLc| (nm)", "Force (pN)", 35)
    if fig is not None: _save(fig, "plot_hexbin_fc_deltalc_abs.png")

    if "Step length [nm]" in df.columns:
        fig = _hexbin_2panel(df["Step length [nm]"], df[force_col],
                             "Unfolding/Refolding – Fc vs Step length (Hexbin)",
                             "Step length (nm)", "Force (pN)", 35)
        if fig is not None: _save(fig, "plot_hexbin_fc_sl.png")

    # Scatter + Jitter
    def _scatter_jitter(ax, x, y, label, color, *,
                        alpha=0.35, size=36,
                        x_jitter_std=0.01, y_jitter_std=0.01,
                        max_points=4000, rng=None):
        x = pd.to_numeric(x, errors="coerce"); y = pd.to_numeric(y, errors="coerce")
        m = (~x.isna()) & (~y.isna()); x, y = x[m].values, y[m].values
        if x.size == 0: return
        if x.size > max_points:
            idx = rng.choice(x.size, size=max_points, replace=False) if rng is not None else np.random.choice(x.size, size=max_points, replace=False)
            x, y = x[idx], y[idx]
        xr = np.ptp(x) if np.ptp(x) > 0 else (abs(np.mean(x)) + 1.0)
        yr = np.ptp(y) if np.ptp(y) > 0 else (abs(np.mean(y)) + 1.0)
        jx = (rng.normal(0, x_jitter_std * xr, size=x.size) if rng is not None else np.random.normal(0, x_jitter_std * xr, size=x.size))
        jy = (rng.normal(0, y_jitter_std * yr, size=y.size) if rng is not None else np.random.normal(0, y_jitter_std * yr, size=y.size))
        ax.scatter(x + jx, y + jy, s=size, alpha=alpha, label=label, edgecolors='none')

    rng = np.random.default_rng(0)
    fig = plt.figure(figsize=(8,5), dpi=120); ax = plt.gca()
    for direction, color, label in zip(['unfolding','refolding'], ['#f5b041','#5dade2'], ['Unfolding','Refolding']):
        sel = (df['Direction']==direction)
        _scatter_jitter(ax, df.loc[sel, "Delta Lc abs (nm)"], df.loc[sel, force_col],
                        label=label, color=color, rng=rng)
    plt.xlabel("|ΔLc| (nm)"); plt.ylabel("Force (pN)")
    plt.title("Unfolding / Refolding – Fc vs |ΔLc| (Scatter + Jitter)")
    plt.legend(); _save(fig, "plot_scatter_jitter_fc_deltalc_abs.png")

    if "Step length [nm]" in df.columns:
        fig = plt.figure(figsize=(8,5), dpi=120); ax = plt.gca()
        for direction, color, label in zip(['unfolding','refolding'], ['#f5b041','#5dade2'], ['Unfolding','Refolding']):
            sel = (df['Direction']==direction)
            _scatter_jitter(ax, df.loc[sel, "Step length [nm]"], df.loc[sel, force_col],
                            label=label, color=color, rng=rng)
        plt.xlabel("Step length (nm)"); plt.ylabel("Force (pN)")
        plt.title("Unfolding / Refolding – Fc vs Step length (Scatter + Jitter)")
        plt.legend(); _save(fig, "plot_scatter_jitter_fc_sl.png")


    # --- Alle CSV FD-Kurven (unfold=blau, refold=rot) + Debug ---
    try:
        # Ordner, in dem die hochgeladene xlsx liegt (== dein smooth Ordner)
        base_dir = str(Path(xlsx_path).parent)
        smooth_csvs = _find_smooth_csvs(base_dir)

        # kleine Debug-Hinweise ins Excel
        debug_lines.append(f"smooth dir searched: {base_dir}")
        debug_lines.append(f"smooth files found: {len(smooth_csvs)}")

        if smooth_csvs:
            fig = plt.figure(figsize=(8, 5), dpi=140)
            ax = plt.gca()
            n_unf, n_ref, n_unk = 0, 0, 0
            skipped = []

            for csvp in smooth_csvs:
                x, y = _read_fd_from_csv(csvp)
                if x is None:
                    skipped.append(Path(csvp).name)
                    smooth_debug_rows.append([Path(csvp).name, "skipped", "", 0, "parser returned None"])
                    continue

                # Richtung aus Dateiname + Pfad
                direction = _direction_from_filename_or_path(Path(csvp).name, csvp)
                if direction == "unfolding":
                    color = "tab:blue"
                    n_unf += 1
                elif direction == "refolding":
                    color = "tab:red"
                    n_ref += 1
                else:
                    color = "0.5"
                    n_unk += 1


                ax.plot(x, y, lw=1.2, alpha=0.5, color=color)
                smooth_debug_rows.append([Path(csvp).name, "plotted", direction, int(len(x)), ""])

            ax.set_title(f"All smooth FD curves (blue=unfold, red=refold) — unf: {n_unf}, ref: {n_ref}, unk: {n_unk}")
            ax.set_xlabel("Distance (nm)"); ax.set_ylabel("Force (pN)")
            ax.set_xlim(left=0); ax.set_ylim(bottom=0); ax.grid(True, alpha=0.2)
            _save(fig, "plot_all_smooth_fd.png")

            if skipped:
                debug_lines.append(f"skipped files: {len(skipped)} (e.g. {', '.join(skipped[:5])}{' ...' if len(skipped)>5 else ''})")
    except Exception as e:
        debug_lines.append(f"[smooth-plot] error: {type(e).__name__}: {e}")

    

    out_mem = io.BytesIO()
    wb = Workbook()

    # --- Data-Sheet ---
    ws_data = wb.active
    ws_data.title = "Data (+ΔLc)"
    for r in dataframe_to_rows(df, index=False, header=True):
        ws_data.append(r)

    # --- Plots-Sheet ---
    ws_plots = wb.create_sheet("Plots")
    row = 1
    # Debug-Zeilen
    for line in debug_lines:
        ws_plots.cell(row=row, column=1, value=line)
        row += 2

    # Bilder einfügen
    for pf in plotfiles:
        try:
            if os.path.exists(pf):
                img = XLImage(pf)
                img.width  = int(img.width * 0.8)
                img.height = int(img.height * 0.8)
                ws_plots.add_image(img, f"A{row}")
                row += max(20, int(img.height / 18))
        except Exception:
            continue

    # --- Smooth-CSV Report (Detailtabelle je Datei) ---
    ws_rep = wb.create_sheet("Smooth-CSV Report")
    for r in smooth_debug_rows:
        ws_rep.append(r)

    wb.save(out_mem)
    out_mem.seek(0)

    # temporäre Plotdateien löschen
    for pf in plotfiles:
        try: os.remove(pf)
        except Exception: pass

    return out_mem

# --------------------------------------------------------------------
# OT-Schema Zielspalten (für manuelle Exporte)
# --------------------------------------------------------------------
OT_COLUMNS = [
    "Filename","step number",
    "Force step start [pN]","Force step end [pN]","mean force [pN]",
    "extension step start [nm]","extension step end [nm]","Step length [nm]",
    "ds contour length","ds persistance Length","ds stiffness (K0) [pN]",
    "ss contour Length","ss persistance Length","ss stiffness (K0) [pN]",
    "Force offset","Distance offset",
    "Work [pN*nm]","Work [kT]","Direction"
]

def _interp_force(FD, distance_nm: float) -> float:
    d = FD[:,1]; f = FD[:,0]
    if d[0] > d[-1]: d = d[::-1]; f = f[::-1]
    return float(np.interp(distance_nm, d, f))

def _mean_force_between(FD, d0: float, d1: float) -> float:
    lo, hi = (d0, d1) if d0 <= d1 else (d1, d0)
    d = FD[:,1]; f = FD[:,0]
    if d[0] > d[-1]: d = d[::-1]; f = f[::-1]
    m = (d >= lo) & (d <= hi); sel = f[m]
    if sel.size == 0:
        return float((_interp_force(FD, lo) + _interp_force(FD, hi))/2.0)
    return float(sel.mean())

def _first_val(x):
    a = np.asarray(x).ravel()
    return float(a[0]) if a.size else float('nan')

def _last_val(x):
    a = np.asarray(x).ravel()
    return float(a[-1]) if a.size else float('nan')

def _build_ot_dataframe_from_tomato_state(token: str) -> pd.DataFrame:
    entry = TOMATO.get(token)
    if not entry or not entry.get("last_steps") or not entry.get("last_results"):
        raise ValueError("Keine Analyse im Speicher. Bitte zuerst analysieren und dann exportieren.")
    filename = entry["filename"]
    path     = entry.get("path", "")
    direction = _direction_from_filename_or_path(filename, path)

    FD = entry["FD"]
    steps = entry["last_steps"]
    results = entry["last_results"]

    ds_row = next((r for r in results if (r.get("step")==0)), None) or {}
    ds_Lc = ds_row.get("Lc_ds", ""); ds_Lp = ds_row.get("Lp_ds", ""); ds_St = ds_row.get("St_ds", "")
    res_by_step = {int(r.get("step")): r for r in results if r.get("step") is not None}

    rows = []
    for i, s in enumerate(steps, start=1):
        start_nm = float(s["start"]); end_nm = float(s["end"]); step_len = abs(end_nm - start_nm)
        f_start = _interp_force(FD, start_nm); f_end = _interp_force(FD, end_nm)
        f_mean  = _mean_force_between(FD, start_nm, end_nm)
        r = res_by_step.get(i, {})
        rows.append({
            "Filename": filename, "step number": i,
            "Force step start [pN]": f_start, "Force step end [pN]": f_end, "mean force [pN]": f_mean,
            "extension step start [nm]": start_nm, "extension step end [nm]": end_nm, "Step length [nm]": step_len,
            "ds contour length": ds_Lc, "ds persistance Length": ds_Lp, "ds stiffness (K0) [pN]": ds_St,
            "ss contour Length": r.get("Lc_ss",""), "ss persistance Length": r.get("Lp_ss",""), "ss stiffness (K0) [pN]": r.get("St_ss",""),
            "Force offset": r.get("f_off",""), "Distance offset": r.get("d_off",""),
            "Work [pN*nm]": r.get("work_pNnm",""), "Work [kT]": r.get("work_kT",""),
            "Direction": direction
        })
    return pd.DataFrame(rows, columns=OT_COLUMNS)

# --------------------------------------------------------------------
# Routes (Startseiten)
# --------------------------------------------------------------------
@app.get("/")
def start():
    return render_template("start.html")

@app.route("/avocado")
def avocado():
    server_files = list_server_files(LAST_SERVER_FOLDER) if LAST_SERVER_FOLDER else []
    return render_template("index.html", server_folder=LAST_SERVER_FOLDER, server_files=server_files, jobs=JOBS)

@app.post("/set_server_folder")
def set_server_folder():
    global LAST_SERVER_FOLDER
    folder = request.form.get("server_folder", "").strip()
    if not folder:
        flash("Bitte einen gültigen Pfad angeben.", "warning")
        return redirect(url_for("avocado"))
    LAST_SERVER_FOLDER = folder
    files = list_server_files(folder)
    flash(f"{len(files)} Datei(en) gefunden.", "info")
    return redirect(url_for("avocado"))

@app.post("/upload_fs")
def upload_fs():
    if "file" not in request.files: return ("no file", 400)
    f = request.files["file"]
    if not f.filename: return ("empty name", 400)
    dst = UPLOAD_DIR / Path(f.filename).name
    f.save(dst)
    return ("ok", 200)

# --------------------------------------------------------------------
# Batch-Run (OT) – Job Handling (unverändert)
# --------------------------------------------------------------------
def run_job(job_id, analysis_folder, ts, files, input_settings, input_format, export_data, input_fitting, q):
    import numpy as _np
    import matplotlib.axes as _maxes
    _orig_fill_between = getattr(_maxes.Axes, 'fill_between')
    def _fill_between_guard(self, x, y1, y2=0, **kw):
        x  = _np.asarray(x); y1 = _np.asarray(y1)
        y2a = _np.asarray(y2) if _np.ndim(y2) else _np.full_like(x[:len(y1)], y2, dtype=float)
        n = min(len(x), len(y1), len(y2a))
        if n <= 0: return
        return _orig_fill_between(self, x[:n], y1[:n], y2a[:n], **kw)
    _maxes.Axes.fill_between = _fill_between_guard
    import POTATO_find_steps as P
    P.save_figure = save_figure_safe
    try:
        start_subprocess(analysis_folder, ts, files, input_settings, input_format, export_data, input_fitting, q)
        q.put("__DONE__")
    except Exception as e:
        q.put(f"__ERROR__:{repr(e)}")

@app.post("/start_run")
def start_run():
    source = request.form.get("source", "server")
    kind = request.form.get("kind", "HF")
    multi = bool(request.form.get("multi"))
    trap1 = bool(request.form.get("trap1"))
    len_um = bool(request.form.get("len_um"))
    defaults = default_values_HF if kind == "HF" else default_values_LF if kind == "LF" else default_values_CSV
    input_settings = make_input_settings(defaults)
    input_format   = make_input_format(kind=kind, trap1=trap1, length_um=len_um, preprocess=True, multi=multi, augment=False)
    model = request.form.get("model", "WLC+WLC")
    model = model if model in ("WLC+WLC", "WLC+FJC") else "WLC+WLC"
    
    # ##################################################################
    # HIER BEGINNT DER ANGEPASSTE CODEBLOCK
    # ##################################################################
    
    # Kopie der Standard-Fit-Parameter erstellen, um sie nicht global zu ändern
    fit_params = default_values_FIT.copy()

    # Auswahl der Persistenzlänge aus dem Formular auslesen
    lp_choice = request.form.get("lp_ds_choice", "default")
    if lp_choice == "50":
        fit_params['Persistance-Length ds, nm'] = '50'
    elif lp_choice == "40":
        fit_params['Persistance-Length ds, nm'] = '40'
    
    # Auswahl der Konturlänge aus dem Formular auslesen
    lc_choice = request.form.get("lc_ds_choice", "default")
    if lc_choice == "1700":
        fit_params['Contour-Length ds, nm'] = '1700' # 1.7 µm
    elif lc_choice == "714":
        fit_params['Contour-Length ds, nm'] = '714'  # 0.714 µm
    elif lc_choice == "3400":
        fit_params['Contour-Length ds, nm'] = '3400' # 3.4 µm

    # Fit-Einstellungen mit den potenziell geänderten Parametern erstellen
    input_fitting = make_input_fitting(fit_params, model=model)
    
    # ##################################################################
    # HIER ENDET DER ANGEPASSTE CODEBLOCK
    # ##################################################################
    
    export_data   = make_export_flags(True, True, True, True, True)

    if source == "server":
        folder = request.form.get("server_folder") or LAST_SERVER_FOLDER
        files = list_server_files(folder)
        if not files:
            flash("Keine .h5/.csv im Server-Ordner gefunden.", "warning")
            return redirect(url_for("avocado"))
    else:
        files = sorted([str(p) for p in UPLOAD_DIR.glob("*") if p.suffix.lower() in (".h5", ".csv")])
        if not files:
            flash("Keine hochgeladenen Dateien gefunden.", "warning")
            return redirect(url_for("avocado"))

    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    analysis_folder = str(ANALYSIS_DIR / f"Analysis_{ts}")
    os.makedirs(analysis_folder, exist_ok=True)
    job_id = f"job-{ts}"

    q = Queue()
    p = Process(target=run_job, args=(job_id, analysis_folder, ts, files, input_settings, input_format, export_data, input_fitting, q))
    p.daemon = True; p.start()

    JOBS[job_id] = {"process": p, "queue": q, "folder": analysis_folder, "ts": ts, "logs": [], "done": False, "error": None, "files": files}
    return redirect(url_for("job_view", job_id=job_id))

@app.get("/job/<job_id>")
def job_view(job_id):
    job = JOBS.get(job_id)
    if not job: abort(404)
    logs = drain_queue(job["queue"])
    for m in logs:
        if m == "__DONE__": job["done"] = True
        elif m.startswith("__ERROR__"): job["error"] = m.replace("__ERROR__:", ""); job["done"] = True
        else: job["logs"].append(m)
    out_files = []
    af = Path(job["folder"])
    if af.exists(): out_files = sorted([p.name for p in af.iterdir() if p.is_file()])
    return render_template("job.html", job_id=job_id, job=job, out_files=out_files)

@app.get("/job/<job_id>/status")
def job_status(job_id):
    job = JOBS.get(job_id)
    if not job: return jsonify({"error": "unknown job"}), 404
    logs = drain_queue(job["queue"]); new_msgs = []
    for m in logs:
        if m == "__DONE__": job["done"] = True
        elif m.startswith("__ERROR__"): job["error"] = m.replace("__ERROR__:", ""); job["done"] = True
        else: job["logs"].append(m); new_msgs.append(m)
    return jsonify({"done": job["done"], "error": job["error"], "new_logs": new_msgs})

@app.get("/job/<job_id>/download/<path:filename>")
def job_download(job_id, filename):
    job = JOBS.get(job_id)
    if not job: abort(404)
    fp = Path(job["folder"]) / filename
    if not fp.exists(): abort(404)
    return send_file(fp, as_attachment=True)

@app.get("/job/<job_id>/download_zip")
def job_download_zip(job_id):
    job = JOBS.get(job_id)
    if not job: abort(404)
    mem = io.BytesIO()
    with zipfile.ZipFile(mem, "w", zipfile.ZIP_DEFLATED) as z:
        for p in Path(job["folder"]).glob("*"):
            if p.is_file(): z.write(p, arcname=p.name)
    mem.seek(0)
    return send_file(mem, as_attachment=True, download_name=f"{job_id}.zip", mimetype="application/zip")
    
# --- PDF-Report direkt aus einem bestehenden Job-Ordner erzeugen ---
from batch_report import generate_batch_pdf  # <— oben bei den Imports ergänzen

@app.get("/job/<job_id>/report.pdf")
def job_report_pdf(job_id):
    job = JOBS.get(job_id)
    if not job:
        abort(404)
    analysis_folder = job.get("folder")
    if not analysis_folder or not Path(analysis_folder).exists():
        abort(404)
    pdf_path, _ = generate_batch_pdf(analysis_folder)  # erzeugt/überschreibt Report im Ordner
    return send_file(pdf_path, as_attachment=True)
    

# --------------------------------------------------------------------
# TOMATO – Seite
# --------------------------------------------------------------------
@app.get("/tomato")
def tomato_page():
    return render_template("tomato.html")

# Datei-Upload
@app.post("/api/tomato/upload")
def tomato_upload():
    if "file" not in request.files:
        return jsonify({"error": "no file"}), 400
    file = request.files["file"]
    if not file.filename:
        return jsonify({"error": "empty filename"}), 400
    dst = (UPLOAD_DIR / file.filename).resolve()
    file.save(dst)

    ext = Path(file.filename).suffix.lower()
    kind_form = request.form.get("kind", "HF")
    if ext == ".csv": kind = "CSV"
    else: kind = "HF" if kind_form not in ("LF", "CSV") else kind_form
    trap1 = request.form.get("trap1", "on") == "on"
    len_um = request.form.get("len_um", "on") == "on"

    input_settings, input_format, input_fitting, export_data = _default_settings(kind, trap1, len_um)
    try:
        FD, FD_um, freq, filename = read_in_data(0, [str(dst)], input_settings, input_format)
    except Exception as e:
        return jsonify({"error": f"read_in_data failed: {type(e).__name__}: {e}"}), 500
    if FD[0, 1] > FD[-1, 1]: FD = np.flipud(FD)

    token = uuid.uuid4().hex
    TOMATO[token] = {
        "filename": filename, "path": str(dst), "FD": FD, "freq": float(freq),
        "settings": input_settings, "format": input_format, "fitcfg": make_input_fitting(default_values_FIT, "WLC+WLC"),
        "last_steps": None, "last_results": None
    }

    x = FD[:, 1]; y = FD[:, 0]
    if len(x) > 6000:
        step = max(1, len(x)//6000); x = x[::step]; y = y[::step]
    return jsonify({"token": token, "filename": filename, "x": x.tolist(), "y": y.tolist(), "n_points": int(len(FD)), "kind": kind})

# Server-Ordner laden
@app.get("/api/tomato/list")
def tomato_list_server_folder():
    folder = request.args.get("folder", "").strip()
    if not folder: return jsonify({"error": "missing folder"}), 400
    files = [f for f in list_server_files(folder) if str(f).lower().endswith((".h5", ".csv"))]
    return jsonify({"files": files})

@app.post("/api/tomato/load_server_file")
def tomato_load_server_file():
    data = request.get_json(force=True, silent=True) or {}
    path = (data.get("path") or "").strip()
    if not path: return jsonify({"error": "missing path"}), 400
    p = Path(path)
    if not p.exists() or not p.is_file(): return jsonify({"error": "file not found"}), 404
    ext = p.suffix.lower()
    if ext == ".csv": kind = "CSV"; trap1 = True; len_um = True
    elif ext == ".h5": kind = "HF"; trap1 = True; len_um = True
    else: return jsonify({"error": f"unsupported extension: {ext}"}), 400

    input_settings, input_format, input_fitting, export_data = _default_settings(kind, trap1, len_um)
    try:
        FD, FD_um, freq, filename = read_in_data(0, [str(p)], input_settings, input_format)
    except Exception as e:
        return jsonify({"error": f"read_in_data failed: {type(e).__name__}: {e}"}), 500
    if FD[0, 1] > FD[-1, 1]: FD = np.flipud(FD)

    token = uuid.uuid4().hex
    TOMATO[token] = {
        "filename": filename, "path": str(p), "FD": FD, "freq": float(freq),
        "settings": input_settings, "format": input_format, "fitcfg": make_input_fitting(default_values_FIT, "WLC+WLC"),
        "last_steps": None, "last_results": None
    }

    x = FD[:, 1]; y = FD[:, 0]
    if len(x) > 6000:
        step = max(1, len(x)//6000); x = x[::step]; y = y[::step]
    return jsonify({"token": token, "filename": filename, "x": x.tolist(), "y": y.tolist(), "n_points": int(len(FD)), "kind": kind})

# --------------------------------------------------------------------
# Manuelle Analyse – (optionale) synchrone Route
# --------------------------------------------------------------------
@app.post("/api/tomato/analyze")
def tomato_analyze():
    data = request.get_json(force=True, silent=True) or {}
    token = data.get("token")
    steps = data.get("steps") or []
    model = data.get("model", "WLC+WLC")
    if not token or token not in TOMATO: return jsonify({"error": "invalid token"}), 400
    if not steps: return jsonify({"error": "no steps provided"}), 400

    entry = TOMATO[token]; FD = entry["FD"]
    input_settings = entry["settings"]; input_fitting = entry["fitcfg"]
    input_fitting = {**input_fitting, "WLC+WLC": 1 if model=="WLC+WLC" else 0, "WLC+FJC": 1 if model=="WLC+FJC" else 0}

    dist_min, dist_max = float(FD[:,1].min()), float(FD[:,1].max())
    results_rows = []; fit_traces = []

    # DS-Fit
    first = steps[0]
    ds_dict, area_ds, ds_end_idx = fitting_ds(
        entry["filename"], input_settings, {"export_FIT":0}, input_fitting,
        float(first["start"]), FD, derivative_array=None, F_low=[], TOMATO_param=1
    )
    distance = np.arange(dist_min, dist_max+2, 2)
    F_ds_model = ds_dict['model_ds'](distance, ds_dict['fit_model'].params)
    fit_traces.append({"name":"DS fit", "x":distance.tolist(), "y":F_ds_model.tolist(), "dash":"dash"})
    results_rows.append({
        "step": 0, "Lc_ds": ds_dict["Lc_ds"], "Lp_ds": ds_dict["Lp_ds"], "St_ds": ds_dict["St_ds"],
        "f_off": ds_dict["f_offset"], "d_off": ds_dict["d_offset"],
        "Lc_ss": "", "Lp_ss": "", "St_ss": "", "work_pNnm": "", "work_kT": ""
    })

    # SS-Segmente
    from POTATO_find_steps import calc_integral
    prev_end = float(first["start"]); prev_area_end = None; prev_d_end = None; prev_f_end = None

    def _fit_ss_segment(d_start, d_end):
        fit_ss, f_fit, d_fit, ss_dict, area_start, area_end = fitting_ss(
            entry["filename"], input_settings, {"export_FIT":0}, input_fitting,
            float(d_start), float(d_end), FD,
            fix=1, max_range=1, derivative_array=None, F_low=[], TOMATO_param=1
        )
        distance = np.arange(dist_min, dist_max+2, 2)
        F_ss_model = ss_dict['model_ss'](distance, fit_ss.params)
        return fit_ss, f_fit, d_fit, ss_dict, area_start, area_end, distance, F_ss_model

    for i in range(1, len(steps)):
        cur = steps[i]
        fit_ss, f_fit, d_fit, ss_dict, area_start, area_end, distance, F_ss_model = _fit_ss_segment(prev_end, float(cur["start"]))
        fit_traces.append({"name": f"SS fit {i}", "x": distance.tolist(), "y": F_ss_model.tolist(), "dash": "dash"})
        if i == 1:
            work_pNnm, kT = calc_integral(area_ds, area_start, prev_end, _first_val(d_fit),
                                          FD[np.searchsorted(FD[:,1], prev_end, side='left')][0], _first_val(f_fit))
        else:
            work_pNnm, kT = calc_integral(prev_area_end, area_start, _last_val(prev_d_end), _first_val(d_fit),
                                          _last_val(prev_f_end), _first_val(f_fit))
        results_rows.append({
            "step": i, "Lc_ds":"", "Lp_ds":"", "St_ds":"", "f_off": ss_dict["f_offset"], "d_off": ss_dict["d_offset"],
            "Lc_ss": ss_dict["Lc_ss"], "Lp_ss": ss_dict["Lp_ss"], "St_ss": ss_dict["St_ss"],
            "work_pNnm": work_pNnm, "work_kT": kT
        })
        prev_end = float(cur["start"]); prev_area_end = area_end; prev_d_end = d_fit; prev_f_end = f_fit

    # letztes Segment
    fit_ss, f_fit, d_fit, ss_dict, area_start, area_end, distance, F_ss_model = _fit_ss_segment(prev_end, float(FD[:,1].max()))
    fit_traces.append({"name": f"SS fit last", "x": distance.tolist(), "y": F_ss_model.tolist(), "dash": "dash"})
    if len(steps) == 1:
        work_pNnm, kT = calc_integral(area_ds, area_start, prev_end, _first_val(d_fit),
                                      FD[np.searchsorted(FD[:,1], prev_end, side='left')][0], _first_val(f_fit))
        results_rows.append({
            "step": 1, "Lc_ds":"", "Lp_ds":"", "St_ds":"",
            "f_off": ss_dict["f_offset"], "d_off": ss_dict["d_offset"],
            "Lc_ss": ss_dict["Lc_ss"], "Lp_ss": ss_dict["Lp_ss"], "St_ss": ss_dict["St_ss"],
            "work_pNnm": work_pNnm, "work_kT": kT
        })
    else:
        results_rows.append({
            "step": len(steps), "Lc_ds":"", "Lp_ds":"", "St_ds":"",
            "f_off": ss_dict["f_offset"], "d_off": ss_dict["d_offset"],
            "Lc_ss": ss_dict["Lc_ss"], "Lp_ss": ss_dict["Lp_ss"], "St_ss": ss_dict["St_ss"],
            "work_pNnm": "", "work_kT": ""
        })

    TOMATO[token]["last_steps"]   = steps
    TOMATO[token]["last_results"] = results_rows
    return jsonify({"filename": entry["filename"], "fits": fit_traces, "results": results_rows})

# --------------------------------------------------------------------
# NEU: Manuelle Analyse als Job (Fortschrittsbalken via Polling, Spawn-fest)
# --------------------------------------------------------------------
def _run_tomato_analyze_worker(entry, steps, model, q):
    """Entry (dict) wird mitgegeben – funktioniert auch im spawn-Modus."""
    import numpy as _np
    from POTATO_find_steps import calc_integral

    def _p(pct, msg): q.put(f"__PROGRESS__:{pct}:{msg}")
    try:
        FD = entry["FD"]
        input_settings = entry["settings"]
        input_fitting  = entry["fitcfg"]
        input_fitting = {**input_fitting,
                         "WLC+WLC": 1 if model=="WLC+WLC" else 0,
                         "WLC+FJC": 1 if model=="WLC+FJC" else 0}

        dist_min, dist_max = float(FD[:,1].min()), float(FD[:,1].max())
        results_rows = []; fit_traces = []

        _p(0, "Starte Analyse…")
        first = steps[0]
        _p(5, "DS-Fit…")
        ds_dict, area_ds, ds_end_idx = fitting_ds(
            entry["filename"], input_settings, {"export_FIT":0}, input_fitting,
            float(first["start"]), FD, derivative_array=None, F_low=[], TOMATO_param=1
        )
        distance = np.arange(dist_min, dist_max+2, 2)
        F_ds_model = ds_dict['model_ds'](distance, ds_dict['fit_model'].params)
        fit_traces.append({"name":"DS fit", "x":distance.tolist(), "y":F_ds_model.tolist(), "dash":"dash"})
        results_rows.append({
            "step": 0, "Lc_ds": ds_dict["Lc_ds"], "Lp_ds": ds_dict["Lp_ds"], "St_ds": ds_dict["St_ds"],
            "f_off": ds_dict["f_offset"], "d_off": ds_dict["d_offset"],
            "Lc_ss": "", "Lp_ss": "", "St_ss": "", "work_pNnm": "", "work_kT": ""
        })
        _p(12, "DS-Fit abgeschlossen.")

        def _fit_ss_segment(d_start, d_end):
            fit_ss, f_fit, d_fit, ss_dict, area_start, area_end = fitting_ss(
                entry["filename"], input_settings, {"export_FIT":0}, input_fitting,
                float(d_start), float(d_end), FD,
                fix=1, max_range=1, derivative_array=None, F_low=[], TOMATO_param=1
            )
            distance = np.arange(dist_min, dist_max+2, 2)
            F_ss_model = ss_dict['model_ss'](distance, fit_ss.params)
            return fit_ss, f_fit, d_fit, ss_dict, area_start, area_end, distance, F_ss_model

        prev_end = float(first["start"]); prev_area_end = None; prev_d_end = None; prev_f_end = None

        for i in range(1, len(steps)):
            _p(12 + int(70*(i-1)/max(1,len(steps)-1)), f"SS-Fit Segment {i}…")
            cur = steps[i]
            fit_ss, f_fit, d_fit, ss_dict, area_start, area_end, distance, F_ss_model = _fit_ss_segment(prev_end, float(cur["start"]))
            fit_traces.append({"name": f"SS fit {i}", "x": distance.tolist(), "y": F_ss_model.tolist(), "dash": "dash"})
            if i == 1:
                work_pNnm, kT = calc_integral(area_ds, area_start, prev_end, float(_first_val(d_fit)),
                                              FD[_np.searchsorted(FD[:,1], prev_end, side='left')][0], float(_first_val(f_fit)))
            else:
                work_pNnm, kT = calc_integral(prev_area_end, area_start, float(_last_val(prev_d_end)), float(_first_val(d_fit)),
                                              float(_last_val(prev_f_end)), float(_first_val(f_fit)))
            results_rows.append({
                "step": i, "Lc_ds":"", "Lp_ds":"", "St_ds":"",
                "f_off": ss_dict["f_offset"], "d_off": ss_dict["d_offset"],
                "Lc_ss": ss_dict["Lc_ss"], "Lp_ss": ss_dict["Lp_ss"], "St_ss": ss_dict["St_ss"],
                "work_pNnm": work_pNnm, "work_kT": kT
            })
            prev_end = float(cur["start"]); prev_area_end = area_end; prev_d_end = d_fit; prev_f_end = f_fit

        _p(85, "SS-Fit letztes Segment…")
        fit_ss, f_fit, d_fit, ss_dict, area_start, area_end, distance, F_ss_model = _fit_ss_segment(prev_end, float(FD[:,1].max()))
        fit_traces.append({"name": f"SS fit last", "x": distance.tolist(), "y": F_ss_model.tolist(), "dash": "dash"})
        if len(steps) == 1:
            work_pNnm, kT = calc_integral(area_ds, area_start, prev_end, float(_first_val(d_fit)),
                                          FD[_np.searchsorted(FD[:,1], prev_end, side='left')][0], float(_first_val(f_fit)))
            results_rows.append({
                "step": 1, "Lc_ds":"", "Lp_ds":"", "St_ds":"",
                "f_off": ss_dict["f_offset"], "d_off": ss_dict["d_offset"],
                "Lc_ss": ss_dict["Lc_ss"], "Lp_ss": ss_dict["Lp_ss"], "St_ss": ss_dict["St_ss"],
                "work_pNnm": work_pNnm, "work_kT": kT
            })
        else:
            results_rows.append({
                "step": len(steps), "Lc_ds":"", "Lp_ds":"", "St_ds":"",
                "f_off": ss_dict["f_offset"], "d_off": ss_dict["d_offset"],
                "Lc_ss": ss_dict["Lc_ss"], "Lp_ss": ss_dict["Lp_ss"], "St_ss": ss_dict["St_ss"],
                "work_pNnm": "", "work_kT": ""
            })

        payload = {"filename": entry["filename"], "fits": fit_traces, "results": results_rows}
        q.put(f"__RESULT__:{json.dumps(payload)}")
        _p(100, "Fertig."); q.put("__DONE__")
    except Exception as e:
        q.put(f"__ERROR__:{repr(e)}")

@app.post("/api/tomato/analyze_start")
def tomato_analyze_start():
    data  = request.get_json(force=True, silent=True) or {}
    token = data.get("token")
    steps = data.get("steps") or []
    model = data.get("model", "WLC+WLC")
    if not token or token not in TOMATO: return jsonify({"error": "invalid token"}), 400
    if not steps: return jsonify({"error": "no steps provided"}), 400

    # Entry für den Subprozess serialisierbar machen (spawn-fest)
    src = TOMATO[token]
    entry_for_worker = {
        "filename": src["filename"],
        "FD":       src["FD"],
        "settings": src["settings"],
        "fitcfg":   src["fitcfg"],
    }

    q = Queue()
    job_id = "toman-" + uuid.uuid4().hex[:8]
    p = Process(target=_run_tomato_analyze_worker, args=(entry_for_worker, steps, model, q))
    p.daemon = True
    p.start()

    TOMATO_JOBS[job_id] = {
        "process": p, "queue": q, "done": False, "error": None, "payload": None,
        "token": token, "steps": steps
    }
    return jsonify({"job_id": job_id})

@app.get("/api/tomato/analyze_status/<job_id>")
def tomato_analyze_status(job_id):
    job = TOMATO_JOBS.get(job_id)
    if not job: return jsonify({"error":"unknown job"}), 404
    progress = _drain_tomato_job(job)
    pct = progress[-1]["pct"] if progress else None
    msg = progress[-1]["msg"] if progress else None
    return jsonify({"done": job["done"], "error": job["error"], "pct": pct, "msg": msg})

@app.get("/api/tomato/analyze_result/<job_id>")
def tomato_analyze_result(job_id):
    job = TOMATO_JOBS.get(job_id)
    if not job: return jsonify({"error":"unknown job"}), 404
    _drain_tomato_job(job)
    if not job["done"]:  return jsonify({"error":"not ready"}), 400
    if job["error"]:     return jsonify({"error": job["error"]}), 500

    payload = job["payload"] or {"fits": [], "results": [], "filename": ""}
    # TOMATO-Session befüllen (für Export-Button usw.)
    token = job.get("token")
    if token in TOMATO and payload and ("results" in payload):
        TOMATO[token]["last_steps"]   = job.get("steps", [])
        TOMATO[token]["last_results"] = payload["results"]
        TOMATO[token]["last_fits"]    = payload.get("fits", [])  # <-- DIESE ZEILE HINZUFÜGEN
    return jsonify(payload)

# --------------------------------------------------------------------
# Export TOMATO (manuelle Analysen) → ZIP (CSV + XLSX)
# --------------------------------------------------------------------
@app.post("/api/tomato/export")
def tomato_export():
    parts = []
    for t, entry in TOMATO.items():
        if entry.get("last_steps") and entry.get("last_results"):
            try: parts.append(_build_ot_dataframe_from_tomato_state(t))
            except Exception: continue
    if not parts:
        return jsonify({"error": "Keine Analysen im Speicher. Bitte erst analysieren."}), 400

    df = pd.concat(parts, ignore_index=True)
    if "Filename" in df.columns and "step number" in df.columns:
        df = df.sort_values(["Filename", "step number"], kind="stable").reset_index(drop=True)

    csv_bytes = df.to_csv(index=False).encode("utf-8-sig")
    xlsx_mem = io.BytesIO()
    with pd.ExcelWriter(xlsx_mem, engine="openpyxl") as xw:
        df.to_excel(xw, index=False, sheet_name="OT-Results")
    xlsx_mem.seek(0)

    zip_mem = io.BytesIO()
    with zipfile.ZipFile(zip_mem, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("tomato_ergebnisse_ot_schema.csv", csv_bytes)
        z.writestr("tomato_ergebnisse_ot_schema.xlsx", xlsx_mem.getvalue())
    zip_mem.seek(0)
    return send_file(zip_mem, as_attachment=True,
                     download_name="tomato_ergebnisse_TOMATO_OT-Ergebnisse.zip",
                     mimetype="application/zip")

# --------------------------------------------------------------------
# NEU: Sammelplot für smooth.csv mit Auswahlliste
# --------------------------------------------------------------------
def find_smooth_files_for_selection(base_dir: str):
    """
    Finds all .csv files containing 'smooth' in their name, recursively.
    """
    p = Path(base_dir)
    if not p.is_dir():
        return []
    return sorted([
        str(f) for f in p.rglob("*.csv")
        if 'smooth' in f.name.lower()
    ])

def categorize_smooth_file(filename: str) -> str:
    """
    Erkennt 'forward'/'reverse' und 'trap'/'ref' aus dem Dateinamen.
    Diese Version ist für komplexe Namen wie '..._3-1f_smooth...' optimiert.
    """
    fn_lower = filename.lower()
    
    # Den Dateinamen an typischen Trennzeichen aufteilen
    parts = re.split(r'[_.-]', fn_lower)
    
    direction = "unknown"

    # Jeden Teil des Namens überprüfen
    for part in parts:
        if not part: continue # Leere Teile überspringen

        # Prüft auf Muster wie "1f", "12f" oder "3-1f"
        if part.endswith('f'):
            base = part[:-1]
            # Prüft, ob der Rest (nach Entfernen von '-') nur aus Ziffern besteht
            if base and base.replace('-', '').isdigit():
                direction = "forward"
                break
        
        # Prüft auf Muster wie "1r", "12r" oder "2-2r"
        if part.endswith('r'):
            base = part[:-1]
            if base and base.replace('-', '').isdigit():
                direction = "reverse"
                break
    
    # Fallback für einfache "fw" oder "rv" Teile
    if direction == "unknown":
        if "fw" in parts:
            direction = "forward"
        elif "rv" in parts:
            direction = "reverse"
    
    # Typ bestimmen (Standard ist "trap")
    file_type = "trap"
    if "ref" in fn_lower:
        file_type = "ref"
    
    if direction == "unknown":
        return "unknown"
        
    return f"{direction}-{file_type}"


@app.post("/api/tomato/list_smooth_files")
def list_smooth_files():
    data = request.get_json(force=True, silent=True) or {}
    folder = data.get("folder", "").strip()
    if not folder or not Path(folder).is_dir():
        return jsonify({"error": "Gültigen Ordner angeben."}), 400
    
    smooth_files = find_smooth_files_for_selection(folder)
    return jsonify({"smooth_files": smooth_files})

@app.post("/api/tomato/plot_selected_smooth")
def plot_selected_smooth():
    data = request.get_json(force=True, silent=True) or {}
    selected_files = data.get("files", [])
    if not selected_files:
        return jsonify({"error": "Keine Dateien ausgewählt."}), 400
    
    curves = []
    for f_path in selected_files:
        try:
            p_file = Path(f_path)
            if not p_file.exists(): continue

            x, y = _read_fd_from_csv(f_path)
            
            if x is not None and y is not None and len(x) > 0:
                # --- NEU: Distanz auf 0 normieren ---
                x_np = np.array(x)
                x_normalized = x_np - x_np.min()
                
                # --- NEU: Präzise Kategorisierung verwenden ---
                category = categorize_smooth_file(p_file.name)

                curves.append({
                    "filename": p_file.name,
                    "x": x_normalized.tolist(),
                    "y": y.tolist(),
                    "category": category,
                })
        except Exception as e:
            print(f"Error processing file {f_path}: {e}")
            continue
            
    return jsonify({"curves": curves})

# --------------------------------------------------------------------
# NEU: Export für Plot-Daten (Rohdaten + Fits als CSV)
# --------------------------------------------------------------------
@app.post("/api/tomato/export_plot_data")
def tomato_export_plot_data():
    data = request.get_json(force=True, silent=True) or {}
    token = data.get("token")
    if not token or token not in TOMATO:
        return jsonify({"error": "Ungültiger Token oder keine Daten geladen."}), 400

    entry = TOMATO[token]
    if "FD" not in entry or entry["FD"].shape[0] == 0:
        return jsonify({"error": "Keine Rohdaten gefunden."}), 400

    # 1. Rohdaten extrahieren
    raw_fd = entry["FD"]
    dist_raw = raw_fd[:, 1]
    force_raw = raw_fd[:, 0]

    # 2. Normalisierung: Kleinsten Distanzwert finden und abziehen
    min_dist = dist_raw.min()
    dist_normalized = dist_raw - min_dist
    
    # Basis-DataFrame mit den normalisierten Rohdaten erstellen
    df_final = pd.DataFrame({
        'Distance_(nm)': dist_normalized,
        'Force_Raw_(pN)': force_raw
    })

    # 3. Fit-Daten interpolieren und hinzufügen
    fits = entry.get("last_fits", [])
    for fit in fits:
        fit_name = fit.get('name', 'fit').replace(' ', '_')
        dist_fit_orig = np.array(fit.get('x', []))
        force_fit_orig = np.array(fit.get('y', []))

        if len(dist_fit_orig) > 1:
            # Wichtig: Auch die Distanz-Achse des Fits normalisieren
            dist_fit_normalized = dist_fit_orig - min_dist
            
            # Kraftwerte des Fits auf die Distanzpunkte der Rohdaten interpolieren
            force_fit_interp = np.interp(dist_normalized, dist_fit_normalized, force_fit_orig)
            
            # Die interpolierte Kurve als neue Spalte hinzufügen
            df_final[f'Force_{fit_name}_(pN)'] = force_fit_interp

    # 4. CSV im Speicher erstellen und für deutschen Excel-Import formatieren
    csv_mem = io.BytesIO()
    df_final.to_csv(csv_mem, index=False, sep=";", decimal=",", float_format='%.4f')
    csv_mem.seek(0)
    
    filename_base = Path(entry.get('filename', 'plot_data')).stem
    download_name = f"{filename_base}_graph_data_aligned.csv"
    
    return send_file(
        csv_mem,
        mimetype='text/csv',
        as_attachment=True,
        download_name=download_name
    )

# --------------------------------------------------------------------
# OT-Auswertung Seite + API
# --------------------------------------------------------------------
@app.get("/ot")
def ot_page():
    return render_template("ot.html")

@app.post("/api/ot/analyze_excel")
def ot_analyze_excel():
    if "file" not in request.files:
        return jsonify({"error": "Keine Datei hochgeladen."}), 400
    f = request.files["file"]
    if not f.filename.lower().endswith(".xlsx"):
        return jsonify({"error": "Bitte eine .xlsx Datei hochladen."}), 400

    tmp_path = (UPLOAD_DIR / Path(f.filename).name).resolve()
    f.save(tmp_path)
    try:
        out_mem = _run_ot_auswertung_on_xlsx(str(tmp_path))
    except Exception as e:
        return jsonify({"error": f"Analyse fehlgeschlagen: {type(e).__name__}: {e}"}), 500
    finally:
        try: os.remove(tmp_path)
        except Exception: pass

    base = Path(f.filename).stem + "_OT-Auswertung_DeltaLC.xlsx"
    return send_file(out_mem, as_attachment=True, download_name=base,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# --------------------------------------------------------------------
# Defaults für Upload-Parsing (Hilfsfunktion)
# --------------------------------------------------------------------
def _default_settings(kind="HF", trap1=True, len_um=True):
    if kind == "HF": cfg = default_values_HF
    elif kind == "LF": cfg = default_values_LF
    else: cfg = default_values_CSV
    s = make_input_settings(cfg)
    f = make_input_format(kind=kind, trap1=trap1, length_um=len_um, preprocess=True, multi=False, augment=False)
    fit = make_input_fitting(default_values_FIT, model="WLC+WLC")
    exp = make_export_flags(export_smooth=False, export_plot=False, export_steps=False, export_total=False, export_fit=False)
    return s, f, fit, exp

# --------------------------------------------------------------------
if __name__ == "__main__":
    # Unter Windows reloader=False setzen, wenn multiprocessing genutzt wird
    app.run(debug=True, use_reloader=False)